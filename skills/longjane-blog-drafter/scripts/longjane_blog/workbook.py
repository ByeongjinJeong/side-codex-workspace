from __future__ import annotations

import json
import os
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Briefing

NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

TOPIC_SHEETS = {
    "숙소": "숙소",
    "맛집·카페": "맛집·카페",
    "관광지": "관광지",
    "교통": "교통",
    "쇼핑·구매": "쇼핑·구매",
    "🏨 숙소": "숙소",
    "🍜 맛집·카페": "맛집·카페",
    "📍 관광지": "관광지",
    "🚄 교통": "교통",
    "🛍 쇼핑·구매": "쇼핑·구매",
}

TOPIC_PREFIXES = {
    "숙소": "STAY",
    "맛집·카페": "FOOD",
    "관광지": "SPOT",
    "교통": "TRAN",
    "쇼핑·구매": "SHOP",
}

TOPIC_TITLE_FIELDS = {
    "숙소": "숙소명",
    "맛집·카페": "식당/카페명",
    "관광지": "관광지명",
    "교통": "이용 수단",
    "쇼핑·구매": "매장명/쇼핑 장소",
}

STATUS_OPTIONS = ["대기중", "작성 시작", "초안 작성 중", "초안 완료", "발행 완료"]
STATUS_MIGRATIONS = {
    "대기 중": "대기중",
    "작성 중": "초안 작성 중",
    "검토 중": "초안 완료",
}

AUTOMATION_COLUMNS = {
    "ID",
    "상태",
    "초안 링크",
    "초안 파일명",
    "메인 키워드",
    "서브 키워드",
    "키워드 판단 근거",
    "키워드 실험 방향",
    "이미지 추가 메모",
    "작성 판단 로그",
    "목표 키워드",
    "블로그탭 노출",
    "인플루언서",
    "인플루언서 점유",
}

REQUIRED_TOPIC_AUTOMATION_COLUMNS = [
    "메인 키워드",
    "서브 키워드",
    "키워드 판단 근거",
    "키워드 실험 방향",
    "이미지 추가 메모",
    "작성 판단 로그",
]

AUTOMATION_PLACEHOLDERS = {
    "메인 키워드": "선택 입력/작성 후 로그",
    "서브 키워드": "선택 입력/작성 후 로그",
    "키워드 판단 근거": "선택 입력/작성 후 로그",
    "키워드 실험 방향": "중복 회피/롱테일 실험 메모",
    "이미지 추가 메모": "작성 후 로그",
    "작성 판단 로그": "작성 후 로그",
}

COMMON_COLUMNS = {
    "ID",
    "방문 날짜",
    "여행지",
    "국내/해외",
    "사진 폴더ID",
    "사진 폴더 ID",
    "목표 키워드",
    "블로그탭 노출",
    "인플루언서",
    "인플루언서 점유",
    "상태",
    "자유 메모",
    "초안 링크",
    "초안 파일명",
}

HEADER_ALIASES = {
    "사진 폴더 ID": "사진 폴더ID",
    "인플루언서 점유": "인플루언서",
    "초안 파일명": "초안 링크",
    "가성비": "가성비 평가",
    "레이오버": "레이오버 여부",
}


class WorkbookReader:
    def __init__(self, path: Path):
        self.path = path
        if not path.exists():
            raise FileNotFoundError(path)
        self._sheets: dict[str, list[list[str]]] | None = None

    def list_briefings(self) -> list[dict[str, str]]:
        rows = []
        seen = set()
        for sheet, topic in TOPIC_SHEETS.items():
            table = self.sheets.get(sheet, [])
            if len(table) < 3:
                continue
            headers = table[1]
            for row_number, values in enumerate(table[3:], start=4):
                row = row_dict(headers, values)
                if not has_article_data(row):
                    continue
                article_id = clean(row.get("ID")) or synthetic_id(topic, row_number, row)
                if not article_id:
                    continue
                seen.add(article_id)
                rows.append({
                    "id": article_id,
                    "topic": topic,
                    "status": clean(row.get("상태")),
                    "title": clean(first_present(row, ["숙소명", "식당/카페명", "관광지명", "이용 수단", "매장명/쇼핑 장소"])),
                })
        status = self.sheets.get("발행 현황", []) or self.sheets.get("📊 발행 현황", [])
        if len(status) >= 4:
            headers = status[1]
            for values in status[3:]:
                row = row_dict(headers, values)
                article_id = clean(row.get("ID"))
                if not article_id or article_id in seen:
                    continue
                rows.append({
                    "id": article_id,
                    "topic": clean(row.get("주제")),
                    "status": clean(row.get("상태")),
                    "title": clean(row.get("제목 (예정)")),
                })
        return rows

    def get_briefing(self, article_id: str) -> Briefing:
        for sheet, topic in TOPIC_SHEETS.items():
            table = self.sheets.get(sheet, [])
            if len(table) < 3:
                continue
            headers = table[1]
            for row_number, values in enumerate(table[3:], start=4):
                row = row_dict(headers, values)
                if not has_article_data(row):
                    continue
                row_id = clean(row.get("ID")) or synthetic_id(topic, row_number, row)
                if row_id != article_id:
                    continue
                row["ID"] = row_id
                if not clean(row.get("사진 폴더ID")):
                    row["사진 폴더ID"] = row_id
                common = {k: clean(v) for k, v in row.items() if k in COMMON_COLUMNS}
                if not common.get("국내/해외"):
                    common["국내/해외"] = infer_region(clean(row.get("여행지")))
                details = {k: clean(v) for k, v in row.items() if k not in COMMON_COLUMNS and clean(v)}
                return Briefing(id=article_id, topic=topic, sheet=sheet, common=common, details=details)
        fallback = self._get_status_briefing(article_id)
        if fallback:
            return fallback
        raise KeyError(f"ID not found: {article_id}")

    @property
    def sheets(self) -> dict[str, list[list[str]]]:
        if self._sheets is None:
            self._sheets = self._read_sheets()
        return self._sheets

    def _read_sheets(self) -> dict[str, list[list[str]]]:
        with zipfile.ZipFile(self.path) as z:
            shared = read_shared_strings(z)
            workbook = ET.fromstring(z.read("xl/workbook.xml"))
            rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
            out = {}
            for sheet in workbook.findall(".//a:sheet", NS):
                name = sheet.attrib["name"]
                rid = sheet.attrib[f"{{{NS['r']}}}id"]
                target = relmap[rid].lstrip("/")
                if not target.startswith("xl/"):
                    target = "xl/" + target
                out[name] = read_sheet(z, target, shared)
            return out

    def _get_status_briefing(self, article_id: str) -> Briefing | None:
        table = self.sheets.get("발행 현황", []) or self.sheets.get("📊 발행 현황", [])
        if len(table) < 4:
            return None
        headers = table[1]
        for values in table[3:]:
            row = row_dict(headers, values)
            if clean(row.get("ID")) != article_id:
                continue
            topic = normalize_topic(clean(row.get("주제")) or infer_topic_from_id(article_id))
            common = {
                "ID": article_id,
                "방문 날짜": clean(row.get("방문 날짜")),
                "여행지": clean(row.get("여행지")),
                "목표 키워드": clean(row.get("목표 키워드")),
                "상태": clean(row.get("상태")),
                "자유 메모": clean(row.get("메모")),
                "초안 링크": clean(row.get("초안 링크")),
            }
            details = {"제목 (예정)": clean(row.get("제목 (예정)"))}
            return Briefing(id=article_id, topic=topic, sheet="📊 발행 현황", common=common, details=details)
        return None


def update_article_status(path: Path, article_id: str, status: str) -> bool:
    if status not in STATUS_OPTIONS:
        raise ValueError(f"Unknown status: {status}")
    if not path.exists():
        raise FileNotFoundError(path)

    keep_vba = path.suffix.lower() == ".xlsm"
    if os.environ.get("LONGJANE_KEEP_WORKBOOK_BACKUP") == "1":
        backup_workbook(path)
    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
    changed = ensure_automation_columns(wb)
    changed = ensure_generated_ids(wb) or changed
    for sheet, topic in TOPIC_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [clean(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)]
        if "상태" not in headers:
            continue
        status_col = headers.index("상태") + 1
        id_col = headers.index("ID") + 1 if "ID" in headers else None
        for row_number in range(4, ws.max_row + 1):
            row_id = clean(ws.cell(row_number, id_col).value) if id_col else ""
            if not row_id or row_id.startswith("="):
                row_values = {
                    headers[col - 1]: ws.cell(row_number, col).value
                    for col in range(1, len(headers) + 1)
                }
                row_id = synthetic_id(topic, row_number, row_values)
            if row_id == article_id:
                if id_col and not clean(ws.cell(row_number, id_col).value):
                    ws.cell(row_number, id_col).value = article_id
                ws.cell(row_number, status_col).value = status
                changed = True
                break
    if changed:
        sync_publication_overview(wb)
        wb.save(path)
    return changed


def update_article_fields(path: Path, article_id: str, fields: dict[str, Any]) -> bool:
    if not path.exists():
        raise FileNotFoundError(path)
    keep_vba = path.suffix.lower() == ".xlsm"
    if os.environ.get("LONGJANE_KEEP_WORKBOOK_BACKUP") == "1":
        backup_workbook(path)
    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
    changed = ensure_automation_columns(wb)
    changed = ensure_generated_ids(wb) or changed
    for sheet, topic in TOPIC_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = sheet_headers(ws)
        id_col = headers.index("ID") + 1 if "ID" in headers else None
        if not id_col:
            continue
        colmap = {header: i + 1 for i, header in enumerate(headers) if header}
        for row_number in range(4, ws.max_row + 1):
            row_id = clean(ws.cell(row_number, id_col).value)
            if not row_id:
                row_id = synthetic_id(topic, row_number, worksheet_row_dict(ws, headers, row_number))
            if row_id != article_id:
                continue
            for key, value in fields.items():
                if key not in colmap:
                    continue
                cell = ws.cell(row_number, colmap[key])
                if clean(cell.value) != clean(value):
                    cell.value = value
                    changed = True
            break
    if changed:
        sync_publication_overview(wb)
        wb.save(path)
    return changed


def sync_workbook(path: Path) -> bool:
    """Write generated IDs and refresh the publication overview sheet."""
    if not path.exists():
        raise FileNotFoundError(path)
    keep_vba = path.suffix.lower() == ".xlsm"
    if os.environ.get("LONGJANE_KEEP_WORKBOOK_BACKUP") == "1":
        backup_workbook(path)
    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
    changed = ensure_automation_columns(wb)
    changed = ensure_generated_ids(wb) or changed
    changed = sync_publication_overview(wb) or changed
    if changed:
        wb.save(path)
    return changed


def ensure_automation_columns(wb) -> bool:
    changed = False
    for sheet in TOPIC_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = sheet_headers(ws)
        for column_name in REQUIRED_TOPIC_AUTOMATION_COLUMNS:
            if column_name in headers:
                continue
            insert_at = len(headers) + 1
            ws.cell(2, insert_at).value = column_name
            ws.cell(3, insert_at).value = AUTOMATION_PLACEHOLDERS.get(column_name, "")
            headers.append(column_name)
            changed = True
    return changed


def ensure_generated_ids(wb) -> bool:
    changed = False
    used: set[str] = set()
    topic_sheets = [(sheet, topic) for sheet, topic in TOPIC_SHEETS.items() if sheet in wb.sheetnames]

    for sheet, _topic in topic_sheets:
        ws = wb[sheet]
        headers = sheet_headers(ws)
        id_col = headers.index("ID") + 1 if "ID" in headers else None
        if not id_col:
            continue
        for row_number in range(4, ws.max_row + 1):
            article_id = clean(ws.cell(row_number, id_col).value)
            if article_id and not article_id.startswith("="):
                used.add(article_id)

    for sheet, topic in topic_sheets:
        ws = wb[sheet]
        headers = sheet_headers(ws)
        id_col = headers.index("ID") + 1 if "ID" in headers else None
        if not id_col:
            continue
        prefix = TOPIC_PREFIXES.get(topic)
        if not prefix:
            continue
        for row_number in range(4, ws.max_row + 1):
            current = clean(ws.cell(row_number, id_col).value)
            if current and not current.startswith("="):
                continue
            row = worksheet_row_dict(ws, headers, row_number)
            if not has_article_data(row):
                continue
            article_id = next_generated_id(prefix, row_number - 3, used)
            ws.cell(row_number, id_col).value = article_id
            used.add(article_id)
            changed = True
    return changed


def sync_publication_overview(wb) -> bool:
    status_sheet = "발행 현황" if "발행 현황" in wb.sheetnames else "📊 발행 현황" if "📊 발행 현황" in wb.sheetnames else ""
    if not status_sheet:
        return False
    ws = wb[status_sheet]
    headers = sheet_headers(ws)
    if "ID" not in headers:
        return False

    colmap = {header: i + 1 for i, header in enumerate(headers) if header}
    if overview_uses_formulas(ws, colmap["ID"]):
        return False
    existing = {}
    for row_number in range(4, ws.max_row + 1):
        article_id = clean(ws.cell(row_number, colmap["ID"]).value)
        if article_id:
            existing[article_id] = row_number

    articles = collect_topic_articles(wb)
    changed = False
    next_row = max(ws.max_row + 1, 4)
    for article in articles:
        article_id = article["ID"]
        row_number = existing.get(article_id)
        if row_number is None:
            row_number = next_row
            next_row += 1
            existing[article_id] = row_number
        values = {
            "ID": article_id,
            "주제": article["주제"],
            "제목(예정)": article["제목"],
            "제목 (예정)": article["제목"],
            "여행지": article["여행지"],
            "방문 날짜": article["방문 날짜"],
            "목표 키워드(자동)": article["키워드"],
            "목표 키워드": article["키워드"],
            "상태": article["상태"],
            "초안 파일명": article["초안 파일명"],
            "메모": article["메모"],
        }
        for header, value in values.items():
            if header not in colmap:
                continue
            cell = ws.cell(row_number, colmap[header])
            if clean(cell.value) != clean(value):
                cell.value = value
                changed = True
    return changed


def overview_uses_formulas(ws: Worksheet, id_col: int) -> bool:
    for row_number in range(4, ws.max_row + 1):
        value = ws.cell(row_number, id_col).value
        if isinstance(value, str) and value.startswith("="):
            return True
    return False


def collect_topic_articles(wb) -> list[dict[str, Any]]:
    articles = []
    for sheet, topic in TOPIC_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = sheet_headers(ws)
        if "ID" not in headers:
            continue
        for row_number in range(4, ws.max_row + 1):
            row = worksheet_row_dict(ws, headers, row_number)
            article_id = clean(row.get("ID"))
            if not article_id or not has_article_data(row):
                continue
            title_field = TOPIC_TITLE_FIELDS.get(topic, "")
            articles.append({
                "ID": article_id,
                "주제": topic,
                "제목": clean(row.get(title_field)),
                "여행지": clean(row.get("여행지")),
                "방문 날짜": row.get("방문 날짜") or "",
                "키워드": clean(first_present(row, ["메인 키워드", "목표 키워드"])),
                "상태": clean(row.get("상태")),
                "초안 파일명": clean(first_present(row, ["초안 파일명", "초안 링크"])),
                "메모": clean(first_present(row, ["작성 판단 로그", "자유 메모"])),
            })
    return articles


def sheet_headers(ws: Worksheet) -> list[str]:
    return [clean(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)]


def worksheet_row_dict(ws: Worksheet, headers: list[str], row_number: int) -> dict[str, Any]:
    return {
        headers[col - 1]: ws.cell(row_number, col).value
        for col in range(1, len(headers) + 1)
        if headers[col - 1]
    }


def has_article_data(row: dict[str, Any]) -> bool:
    return any(clean(value) for key, value in row.items() if key not in AUTOMATION_COLUMNS)


def next_generated_id(prefix: str, preferred_number: int, used: set[str]) -> str:
    candidate = f"{prefix}-{preferred_number:03d}"
    if candidate not in used:
        return candidate
    n = 1
    pattern = re.compile(rf"^{re.escape(prefix)}-(\d+)$")
    for article_id in used:
        match = pattern.match(article_id)
        if match:
            n = max(n, int(match.group(1)) + 1)
    while f"{prefix}-{n:03d}" in used:
        n += 1
    return f"{prefix}-{n:03d}"


def backup_workbook(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}_backup_{timestamp}{path.suffix}")
    if not backup.exists():
        shutil.copy2(path, backup)
    return backup


def read_shared_strings(z: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in z.namelist():
        return []
    root = ET.fromstring(z.read("xl/sharedStrings.xml"))
    values = []
    for si in root.findall("a:si", NS):
        values.append("".join(t.text or "" for t in si.findall(".//a:t", NS)))
    return values


def read_sheet(z: zipfile.ZipFile, target: str, shared: list[str]) -> list[list[str]]:
    root = ET.fromstring(z.read(target))
    rows: list[list[str]] = []
    for row in root.findall(".//a:sheetData/a:row", NS):
        cells: dict[int, str] = {}
        for cell in row.findall("a:c", NS):
            ref = cell.attrib.get("r", "")
            col = col_index(ref)
            cells[col] = cell_value(cell, shared)
        if not cells:
            rows.append([])
            continue
        width = max(cells) + 1
        rows.append([cells.get(i, "") for i in range(width)])
    return rows


def cell_value(cell: ET.Element, shared: list[str]) -> str:
    typ = cell.attrib.get("t")
    if typ == "inlineStr":
        return "".join(t.text or "" for t in cell.findall(".//a:t", NS))
    value = cell.find("a:v", NS)
    if value is None or value.text is None:
        return ""
    text = value.text
    if typ == "s":
        try:
            return shared[int(text)]
        except (ValueError, IndexError):
            return text
    return text


def col_index(ref: str) -> int:
    match = re.match(r"([A-Z]+)", ref)
    if not match:
        return 0
    idx = 0
    for ch in match.group(1):
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def row_dict(headers: list[str], values: list[str]) -> dict[str, Any]:
    row = {}
    for i, header in enumerate(headers):
        key = clean(header)
        if not key:
            continue
        key = HEADER_ALIASES.get(key, key)
        row[key] = values[i] if i < len(values) else ""
    return row


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text == "None" else text


def first_present(row: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        if clean(row.get(key)):
            return clean(row.get(key))
    return ""


def infer_topic_from_id(article_id: str) -> str:
    prefix = article_id.split("-", 1)[0].upper()
    return {
        "STAY": "숙소",
        "FOOD": "맛집·카페",
        "SPOT": "관광지",
        "TRANSIT": "교통",
        "TRAN": "교통",
        "SHOP": "쇼핑·구매",
    }.get(prefix, "")


def normalize_topic(topic: str) -> str:
    return {
        "맛집": "맛집·카페",
        "카페": "맛집·카페",
        "쇼핑": "쇼핑·구매",
        "구매": "쇼핑·구매",
        "쇼핑구매": "쇼핑·구매",
    }.get(topic, topic)


def infer_region(destination: str) -> str:
    if not destination:
        return ""
    domestic_markers = [
        "서울",
        "부산",
        "제주",
        "인천",
        "대구",
        "대전",
        "광주",
        "울산",
        "세종",
        "경기",
        "강원",
        "충북",
        "충남",
        "전북",
        "전남",
        "경북",
        "경남",
        "대한민국",
        "한국",
    ]
    return "국내" if any(marker in destination for marker in domestic_markers) else "해외"


def synthetic_id(topic: str, row_number: int, row: dict[str, Any]) -> str:
    if not clean(row.get("방문 날짜")):
        return ""
    prefix = TOPIC_PREFIXES.get(topic)
    if not prefix:
        return ""
    return f"{prefix}-{row_number - 3:03d}"
